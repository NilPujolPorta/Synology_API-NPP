import requests
import json
from datetime import datetime,timezone
import datetime
from os.path import exists	
from openpyxl import Workbook
from openpyxl import load_workbook
import os
import argparse
import mysql.connector
import yaml
from tqdm import tqdm

#millores a fer:
# apendre a anexar al excel per poder fer algunes funcions actualment commentades

# Escriure o llegir del fitxer de config/config.yaml, en el qual es guarda la ultima data on es va agafar dades de synology menys un mes
# WoR determina si escriu "w" o si llegeix "r"
# si es tria l'opcio de llegir retorna un string amb el temps en utc timestamp, sino no retorna res 
def Data(WoR):
	if WoR == "w":
		with open("config/config.yaml") as f:
			list_doc = yaml.safe_load(f)
		list_doc[0]['data']=str(temps()-2592000)
		with open("config/config.yaml", 'w') as yamlfilew:
			yaml.dump(list_doc, yamlfilew)
			
	elif WoR == "r":
		with open("config/config.yaml", 'r') as yamlfiler:
			data = yaml.load(yamlfiler, Loader=yaml.FullLoader)
			return(data[0]['data'])
	else:
		print("Error en modificar config/config.yaml (el metode de interaccio amb el fitxer es erroni o inexistent)")

#Retorna el temps actual en utc timestamp
def temps():
	dt = datetime.datetime.now(timezone.utc)
	utc_time = dt.replace(tzinfo=timezone.utc)
	utc_timestamp = utc_time.timestamp()
	return(round(utc_timestamp))

#Es logueja en la webapi de synology 
#Els parametres son les credencials, la url per fer el logeig i la cookie identificacio enlloc de la sid.
#Retorna la sid que servira per identificar-nos en les operacions seguents
def login(user, password, url, cookie):
	login_parameters = {"api":"SYNO.API.Auth", "version":"3", "method":"login", "account": user, "passwd": password, "session":"ActiveBackup", "format":"cookie"}
	my_headers = {"cookie": cookie}
	response = requests.get(url, params=login_parameters, headers=my_headers).json()
	sid = response['data']['sid']
	
	if	response['success'] == True:
		if args.quiet:
			print("Login correcte")
	else:
		print("Login erroni")
		now = datetime.datetime.now()
		date_string = now.strftime('%Y-%m-%d--%H-%M-%S-login')
		f = open("errorLogs/"+date_string+".txt",'w')
		f.write(str(response))
		f.close()
		print(response)
	return(sid)

#Tanca la sessió anteriorment oberta per la funcio login
#Els paramatres es la url del lloc de logout i la sid i la cookie per idenficació
#Retorna la resposta de la webapi tan si es error com si es correcte
def logout(url, sid, cookie):
	logout_parameters = {"api":"SYNO.API.Auth", "version":"2", "method":"logout", "session":"ActiveBackup"}
	my_headers={"cookie": cookie}
	response = requests.get(url, params=logout_parameters, headers=my_headers).json()
	if	response['success'] == True:
		if args.quiet:
			print("Logout correcte")
		return(response)
	else:
		print("Logout erroni")
		now = datetime.datetime.now()
		date_string = now.strftime('%Y-%m-%d--%H-%M-%S-Logout')
		f = open("errorLogs/"+date_string+".txt",'w')
		f.write(str(response))
		f.close()
		print(response)

#Aconsegueix la informacio de les copies de seguretat de un NAS
#Els parametres son la sid i la cookie per identificació i la url del NAS al cual recolectar les dades
#Retorna les dades en format json i en cas de que dongui error retorna un text en format json sense dades per aixis evitar el issue #3 "Error en les dades que retorna despres de que es trobi amb un nas sense connexio"
def InfoCopies(url, cookie, sid):#6 issue. A vegades dona error sense motiu aparent al fer-ho una segona vega es soluciona
	copies_parameters = {"api":"SYNO.ActiveBackup.Overview", "version":"1", "method":"list_device_transfer_size", "time_start": int(Data("r")), "time_end": temps(), "_sid": sid}
	response = requests.get(url, params=copies_parameters, headers={"cookie":cookie}).json()
	if	response['success'] == True:
		if args.quiet:
			print("Operacio de dades de backup correcte")
		return(response)
	else:
		print("Operacio de dades de backup erroni")
		now = datetime.datetime.now()
		date_string = now.strftime('%Y-%m-%d--%H-%M-%S-Backups')
		f = open("errorLogs/"+date_string+".txt",'w')
		f.write(str(response))
		f.close()
		print(response)
		return({'data': {'device_list': [], 'total': 0}, 'success': False})

#Recull totes les dades de tots els NAS
#El paramatre workbook es per si la opció del excel es activa escriu dades el excel si no es pot connectar amb la maquina
#Retorna un array de text en format json amb les dades de cada NAS
def recoleccioDades(workbook):
	global current_transaction
	global fitxer
	Backups = []
	for x in taulabd:
		#aconseguir un apartat de la url (aut.cgi o entry.cgi)
		cookie = x[4]
		query_parameters = {"api":"SYNO.API.Info", "version":"1", "method":"query", "query":"all"}
		queryUrl = x[3]+"webapi/query.cgi"
		try:
			query = requests.get(queryUrl, params=query_parameters, headers={"cookie": cookie}).json()
			path= str(query['data']['SYNO.API.Auth']['path'])
		except:
			print("Error en la query")
			now = datetime.datetime.now()
			date_string = now.strftime('%Y-%m-%d--%H-%M-%S-Query')
			a = open("errorLogs/"+date_string+".txt",'w')
			a.write(str(query))
			a.close()
		user = x[1]
		password = x[2]
		url = x[3]+"webapi/"+path
		url2 = x[3]+"webapi/entry.cgi"
		nom = x[0]
		if args.quiet:
			print(nom)
		try:
			sid = login(user, password, url, cookie)
			Backups.append(InfoCopies(url2, cookie, sid))
			logout(url, sid, cookie)
		except:
			now = datetime.datetime.now()
			date_string = now.strftime('%Y-%m-%d--%H-%M-%S-Conexio')
			f = open("errorLogs/"+date_string+".txt",'w')
			f.write("Error en connectar amb la maquina "+nom)
			f.close()
			print("Error en connectar amb la maquina")
			if args.excel:
				wsdefault = workbook['Sheet']
				wsdefault.cell(row=current_transaction, column=1, value=nom)
				wsdefault.cell(row=current_transaction, column=2, value="-")
				wsdefault.cell(row=current_transaction, column=3, value="Error en connectar amb la maquina")
				wsdefault.cell(row=current_transaction, column=6, value="-")
				workbook.save(fitxer)
			current_transaction += 1
			codenaError = {'data': {'total': 0}, 'success': False}
			Backups.append(codenaError)
		if args.quiet:
			print()
	Data("w")	#escriure la ultima data aixis sap desde on mirar les copies, per activar aixo primer he de fer que anexi a el fitxer on envia
	return(Backups)

#Interpreta el codi de status de les copies
#El parametre es el codi(int) que ha donat la webapi
#Retorna el estatus(str) en el cual es troba la copia de seguretat
def statusConvertor(status):
	if status == 2:
		return("Correcte")
	elif status == 4:
		return("Warning")
	elif status == 5:
		return("ERROR")
	else:
		return("codi desconegut")

#Aconsegueix el tamany lliure que li queda a un NAS
#El parametre es el numero del NAS el ordre depen de la base de dades
#Retorna el tamany en GB si tot surt be i si no retorna el error
def tamanyRestant(i):
	url2 = taulabd[i][3]+"webapi/entry.cgi"
	nom = taulabd[i][0]
	cookie = taulabd[i][4]
	#aconseguir un apartat de la url (aut.cgi o entry.cgi)
	queryUrl = taulabd[i][3]+"webapi/query.cgi"
	query_parameters = {"api":"SYNO.API.Info", "version":"1", "method":"query", "query":"all"}
	try:
		query = requests.get(queryUrl, params=query_parameters, headers={"cookie": cookie}).json()
		path= str(query['data']['SYNO.API.Auth']['path'])
		url = taulabd[i][3]+"webapi/"+path
	except Exception as e:
		now = datetime.datetime.now()
		date_string = now.strftime('%Y-%m-%d--%H-%M-%S-Query')
		a = open("errorLogs/"+date_string+".txt",'w')
		a.write(str(e))
		a.close()
	user = taulabd[i][1]
	password = taulabd[i][2]
	my_headers = {"cookie": cookie}
	try:
		sid = login(user, password, url, cookie)
		tamany_parameters = {"api":"SYNO.FileStation.List", "version":"2", "method":"list_share", "additional":'["volume_status"]', "_sid": sid}
		response = requests.get(url2, params=tamany_parameters, headers=my_headers).json()
		espaiLliure = round(((response['data']['shares'][0]['additional']['volume_status']['freespace']/1024)/1024)/1024)
		return(espaiLliure)
	except Exception as e:
		now = datetime.datetime.now()
		date_string = now.strftime('%Y-%m-%d--%H-%M-%S-Conexio')
		f = open("errorLogs/"+date_string+".txt",'w')
		f.write("Error en connectar amb la maquina "+ nom)
		f.close()
		return("Fallo en la conexio")

# y es cada transaccio (es reseteja per cada dispositiu)
# z es personalitzat que es per cada dispositiu que tingui transaccio (es reseteja per cada NAS)
# x es cada dispositiu (es reseteja per cada NAS)
# i es cada nas (es reseteja cada execucio)
# current_transaction es cada transaccio (es reseteja cada execucio)
#Escriu les dades en un excel, Nomes s'executa quan l'opcio de l'excel esta activada
##Els parametres inclouen la fulla de excel a on ho escriura(ws), les dades que escriura, i a quines files i columnes (y, z)
#No retorna res
def escriureDades(nom_dispositiu, ws, status, temps_finalitzacio, tamany_transferencia, y, z, nom_nas, tamanyLliure):
	global current_transaction
	file_time = datetime.datetime.fromtimestamp(temps_finalitzacio)
	num1=4*z+1
	num2=4*z+2
	num3=4*z+3

	d = ws.cell(row=num1, column=1, value=nom_dispositiu)
	d = ws.cell(row=num2, column=1, value="Status")
	d = ws.cell(row=num3, column=1, value="Tamany MB")
	d = ws.cell(row=num1, column=y+2, value=file_time.strftime('%Y-%m-%d'))
	d = ws.cell(row=num2, column=y+2, value=status)
	d = ws.cell(row=num3, column=y+2, value=round((tamany_transferencia/1024)/1024))

	wsdefault = workbook['Sheet']

	if y==0:
		
		wsdefault.cell(row=current_transaction, column=1, value=nom_nas)
		wsdefault.cell(row=current_transaction, column=2, value=nom_dispositiu)
		wsdefault.cell(row=current_transaction, column=3, value=('=LOOKUP(2,1/('+nom_nas+'!'+str(num1)+':'+str(num1)+'<>""),'+nom_nas+'!'+str(num1)+':'+str(num1)+')'))
		wsdefault.cell(row=current_transaction, column=4, value=('=LOOKUP(2,1/('+nom_nas+'!'+str(num3)+':'+str(num3)+'<>""),'+nom_nas+'!'+str(num3)+':'+str(num3)+')'))
		wsdefault.cell(row=current_transaction, column=5, value=('=LOOKUP(2,1/('+nom_nas+'!'+str(num2)+':'+str(num2)+'<>""),'+nom_nas+'!'+str(num2)+':'+str(num2)+')'))
		wsdefault.cell(row=current_transaction, column=6, value=tamanyLliure)
		current_transaction += 1

#Gestiona l'escriptura a l'excel i a quina fulla escriu i comprova si la fulla existeix o no (lo ultim ja no faria falta ja que les borro totes al principi)
#Els parametres son: el excel(workbook), les files i columnes a on s'escriura(y, z) i les dades
#No retorna res
def escriptorExcel(nom_dispositiu, status, temps_finalitzacio, tamany_transferencia, workbook, y, z, nom_nas, tamanyLliure):
	trobat = False
	for sheet in workbook:
		if sheet.title == nom_nas:
			trobat = True
			escriureDades(nom_dispositiu, sheet, status, temps_finalitzacio, tamany_transferencia, y, z, nom_nas, tamanyLliure)
			break
	if trobat == False:
		ws = workbook.create_sheet(nom_nas)
		escriureDades(nom_dispositiu, ws, status, temps_finalitzacio, tamany_transferencia, y, z, nom_nas, tamanyLliure)

def borrar(sheet, row):
	for cell in row:
		if cell.value == None:
			return
		sheet.delete_rows(row[0].row, 1)

#Prepara la fulla principal del excel en cas de que no existis abans
#L'únic parametre es el document d'excel
#No retorna res. S'hauria de fer que tambe dones format condicional entre altres
def prepExcel(workbook):
	if args.quiet:
		print("Preparant excel")
	for sheet in workbook:
		if sheet.title != "Sheet":
			workbook.remove(sheet)
			
	wsdefault = workbook['Sheet']
	
	for row in wsdefault:
		borrar(wsdefault,row)			

	wsdefault.cell(row=1, column=1, value="Nom NAS")
	wsdefault.cell(row=1, column=2, value="Nom Dispositiu")
	wsdefault.cell(row=1, column=3, value="Data")
	wsdefault.cell(row=1, column=4, value="Tamany MB")
	wsdefault.cell(row=1, column=5, value="Status")
	wsdefault.cell(row=1, column=6, value="Tamany Lliure GB")


#Acces a la base de dades i recoleccio de la informacio
#Els parametres son les credencials i la ip/host de la base de dades
#Retorna una llista igual a la base de dades
def bd(servidor, usuari, contrassenya):
	try:
		mydb =mysql.connector.connect(
    	    host=servidor,
    	    user=usuari,
    	    password=contrassenya,
    	    database="synology"
    	    )
		mycursor = mydb.cursor(buffered=True)
		if args.quiet:
			print("Access BDD correcte")
	except:
		try:        
			mydb =mysql.connector.connect(
	            host=servidor,
	            user=usuari,
	            password=contrassenya
	            )
			print("Base de dades no existeix, creant-la ...")
			mycursor = mydb.cursor(buffered=True)
			mycursor.execute("CREATE DATABASE synology")
			mydb =mysql.connector.connect(
	            host=servidor,
	            user=usuari,
	            password=contrassenya,
	            database="synology"
	            )
			mycursor = mydb.cursor(buffered=True)
			mycursor.execute("CREATE TABLE dispositius (nom VARCHAR(255), usuari VARCHAR(255), contassenya VARCHAR(255), url VARCHAR(255), cookie VARCHAR(400), pandoraID INT(3));")
		except:
			print("Login BDD incorrecte")
			quit()
	taulabdi = []

	mycursor.execute("SELECT * FROM dispositius")
	resultatbd = mycursor.fetchall()
	for fila in resultatbd:
		taulabdi.append(fila)
	return(taulabdi)

#Escriu les dades finals en un fitxer .json
#El paramatra es la llista a on estan guardades les dades tot i que no faria falta posarla com a parametre ja que es global
#No retorna res
def escriureDadesJSON(llistaFinal):
	if exists("dadesSynology.json") == True:
			os.remove("dadesSynology.json")
	try:
		with open("dadesSynology.json", 'w') as f:
			json.dump(llistaFinal, f, indent = 4)
	except Exception as e:
			print("Error d'escriptura de json")
			now = datetime.datetime.now()
			date_string = now.strftime('%Y-%m-%d--%H-%M-%S-json')
			f = open("errorLogs/"+date_string+".txt",'w')
			f.write("Error d'escriptura de json "+str(e))
			f.close()
	if not(args.quiet):
		print("Done")

###################################################################################################################################################################

parser = argparse.ArgumentParser(description='Una API per a recullir invormacio de varis NAS Synology que tinguin la versio 6 o mes.', epilog="Per configuracio adicional anar a config/config.yaml")
parser.add_argument('-e', '--excel', help='Guardar la informacio a un excel, per defecte esta desactivat', action="store_true")
parser.add_argument('-q', '--quiet', help='Nomes mostra els errors i el missatge de acabada per pantalla.', action="store_false")
parser.add_argument('-f', '--file', help='Especificar el fitxer de excel a on guardar. Per defecte es revisio_copies_seguretat_synology_vs1.xlsx', default="revisio_copies_seguretat_synology_vs1.xlsx", metavar="RUTA")
parser.add_argument('-v', '--versio', help='Mostra la versio', action='version', version='Synology_API-NPP vs1.6.5')
args = parser.parse_args()

current_transaction = 2
fitxer = args.file 

if not(exists("config/config.yaml")):
	print("Emplena el fitxer de configuracio de Base de Dades a config/config.yaml")
	article_info = [
    	{
        	'BD': {
    	    'host' : 'localhost',
    	    'user': 'root',
    	    'passwd': 'patata'
    	    },
			'data': str(temps()-2592000)
    	}
	]
	with open("config/config.yaml", 'w') as yamlfile:
		data = yaml.dump(article_info, yamlfile)

with open("config/config.yaml", "r") as yamlfile:
    data = yaml.load(yamlfile, Loader=yaml.FullLoader)

servidor = data[0]['BD']['host']
usuari = data[0]['BD']['user']
contrassenya = data[0]['BD']['passwd']

taulabd = bd(servidor, usuari, contrassenya)

if exists(fitxer) == False:
	workbook = Workbook()
	prepExcel(workbook)
	workbook.save(fitxer)
elif args.excel:
	workbook = load_workbook(filename = fitxer)
	prepExcel(workbook)

else:
	workbook = load_workbook(filename = fitxer)


llistaTransf = []
llistadispCopia = []
llistaNAS = []
dadesCopiesTotes = recoleccioDades(workbook)
num_nas = len(dadesCopiesTotes)

# y es cada transaccio (es reseteja per cada dispositiu)
# z es personalitzat que es per cada dispositiu que tingui transaccio (es reseteja per cada NAS)
# x es cada dispositiu (es reseteja per cada NAS)
# nas es cada nas (es reseteja cada execucio)
# current_transaction es cada transaccio (es reseteja cada execucio)
nas = 0
nom_dispositiu=""
for nas in tqdm (range(num_nas), desc="Processar Dades", ncols=100, disable=args.quiet):
	nom_nas = taulabd[nas][0]
	id_pandora = taulabd[nas][5]
	num_copies = int(dadesCopiesTotes[nas]['data']['total'])
	tamanyLliure=tamanyRestant(nas)
	x = 0
	z = 0
	while x < num_copies: 
		if len(dadesCopiesTotes[nas]['data']['device_list'][x]['transfer_list']) != 0:
			num_transferencies = len(dadesCopiesTotes[nas]['data']['device_list'][x]['transfer_list'])

			y=0
			for y in tqdm (range (num_transferencies), desc=nom_nas +" | "+ dadesCopiesTotes[nas]['data']['device_list'][x]['transfer_list'][y]['device_name'], ncols=125, disable=not(args.quiet)):
				nom_dispositiu = dadesCopiesTotes[nas]['data']['device_list'][x]['transfer_list'][y]['device_name']
				status = statusConvertor(dadesCopiesTotes[nas]['data']['device_list'][x]['transfer_list'][y]['status'])
				tamany_transferencia = dadesCopiesTotes[nas]['data']['device_list'][x]['transfer_list'][y]['transfered_bytes']
				temps_finalitzacio = dadesCopiesTotes[nas]['data']['device_list'][x]['transfer_list'][y]['time_end']

				file_time = datetime.datetime.fromtimestamp(temps_finalitzacio)
				dataF=file_time.strftime('%Y-%m-%d')
				if y+1 < num_transferencies:
					u=1
				else:
					llistaTransf.append({"data":dataF, "status":status, "tamany_transferenciaMB":(tamany_transferencia/1024)/1024})
				if args.excel:
					escriptorExcel(nom_dispositiu, status, temps_finalitzacio, tamany_transferencia, workbook, y, z, nom_nas, tamanyLliure)
					try:
						workbook.save(fitxer)
					except:
						now = datetime.datetime.now()
						date_string = now.strftime('%Y-%m-%d--%H-%M-%S-permisos')
						f = open("errorLogs/"+date_string+".txt",'w')
						f.write("Error de permisos en obrir el Excel (Pot ser que el excel estigui obert?)")
						f.close()
						print("Error de permisos")
			z += 1
		llistadispCopia.append({"nomDispositiu":nom_dispositiu, "Transferencies":llistaTransf})
		nom_dispositiu = ""
		llistaTransf = []
		x += 1
	llistaNAS.append({"nomNAS":nom_nas, "ID Pandora":id_pandora, "copies":llistadispCopia})
	llistadispCopia = []
	nas +=1
escriureDadesJSON([{"NAS":llistaNAS}])