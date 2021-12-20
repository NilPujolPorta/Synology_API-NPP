import requests
import json
from datetime import datetime,timezone
import datetime
import time
from os.path import exists	
import openpyxl				#pip install openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import os

#millores a fer:
# apendre a anexar al excel per poder fer algunes funcions actualment commentades

current_transaction = 2
fitxer = "revisio_copies_seguretat_synology_vs0.4.xlsx" #sobra pero es posa per si de cas no funciones el de sota

f = open("config/api.conf",'r')
linea = str(f.readlines())
llargada = len(linea)-2
linea = linea[8:llargada]
fitxer = linea
f.close()

def Data(WoR):
	if WoR == "w":
		with open('data/data.txt', 'w') as f:
			wdata = temps()-2592000
			f.write(str(wdata))
			f.close()
	elif WoR == "r":
		with open('data/data.txt', 'r') as f:
			rdata = f.read()
			f.close()
			return(rdata)
	else:
		print("Error en modificar data/data.txt (el metode de interaccio amb el fitxer es erroni o inexistent)")

def temps():
	dt = datetime.datetime.now(timezone.utc)
	utc_time = dt.replace(tzinfo=timezone.utc)
	utc_timestamp = utc_time.timestamp()
	return(round(utc_timestamp))

def login(user, password, url, cookie):
	login_parameters = {"api":"SYNO.API.Auth", "version":"3", "method":"login", "account": user, "passwd": password, "session":"ActiveBackup", "format":"cookie"}
	my_headers = {"cookie": cookie}
	response = requests.get(url, params=login_parameters, headers=my_headers).json()
	sid = response['data']['sid']
	
	if	response['success'] == True:
		print("Login correcte")
	else:
		print("Login erroni")
		now = datetime.datetime.now()
		date_string = now.strftime('%Y-%m-%d--%H-%M-%S-login')
		f = open("errorLogs/"+date_string+".txt",'w')
		f.write(str(response))
		f.close()
		print(response)
	return(sid)

def logout(url, sid, cookie):
	logout_parameters = {"api":"SYNO.API.Auth", "version":"2", "method":"logout", "session":"ActiveBackup"}
	my_headers={"cookie": cookie}
	response = requests.get(url, params=logout_parameters, headers=my_headers).json()
	if	response['success'] == True:
		print("Logout correcte")
		return(response)
	else:
		print("Logout erroni")
		now = datetime.datetime.now()
		date_string = now.strftime('%Y-%m-%d--%H-%M-%S-Logout')
		f = open("errorLogs/"+date_string+".txt",'w')
		f.write(str(response))
		f.close()
		print(response)

def InfoCopies(url, cookie, sid):
	copies_parameters = {"api":"SYNO.ActiveBackup.Overview", "version":"1", "method":"list_device_transfer_size", "time_start": int(Data("r")), "time_end": temps(), "_sid": sid}
	response = requests.get(url, params=copies_parameters, headers={"cookie":cookie}).json()
	if	response['success'] == True:
		print("Operacio de dades de backup correcte")
		return(response)
	else:
		print("Operacio de dades de backup erroni")
		now = datetime.datetime.now()
		date_string = now.strftime('%Y-%m-%d--%H-%M-%S-Backups')
		f = open("errorLogs/"+date_string+".txt",'w')
		f.write(str(response))
		f.close()
		print(response)

def recoleccioDades(workbook):
	global current_transaction
	global fitxer
	with open('data/dispositius.json', 'r') as f:
		rconf = json.load(f)
		num_dispositius = len(rconf['dispositius'])
		Backups = []

		i=0
		while i < num_dispositius:
			#aconseguir un apartat de la url (aut.cgi o entry.cgi)
			cookie = rconf['dispositius'][i]['cookie']
			query_parameters = {"api":"SYNO.API.Info", "version":"1", "method":"query", "query":"all"}
			queryUrl = rconf['dispositius'][i]['url']+"webapi/query.cgi"
			try:
				query = requests.get(queryUrl, params=query_parameters, headers={"cookie": cookie}).json()
				path= str(query['data']['SYNO.API.Auth']['path'])
			except:
				print("Error en la query")
				now = datetime.datetime.now()
				date_string = now.strftime('%Y-%m-%d--%H-%M-%S-Query')
				a = open("errorLogs/"+date_string+".txt",'w')
				a.write(str(query))
				a.close()

			user = rconf['dispositius'][i]['user']
			password = rconf['dispositius'][i]['password']
			url = rconf['dispositius'][i]['url']+"webapi/"+path
			url2 = rconf['dispositius'][i]['url']+"webapi/entry.cgi"
			nom = rconf['dispositius'][i]['nom']
			print(nom)

			try:
				sid = login(user, password, url, cookie)
				Backups.append(InfoCopies(url2, cookie, sid))
				logout(url, sid, cookie)
			except:
				now = datetime.datetime.now()
				date_string = now.strftime('%Y-%m-%d--%H-%M-%S-Conexio')
				f = open("errorLogs/"+date_string+".txt",'w')
				f.write("Error en connectar amb la maquina "+nom)
				f.close()
				print("Error en connectar amb la maquina")
				wsdefault = workbook['Sheet']
				wsdefault.cell(row=current_transaction, column=1, value=nom)
				wsdefault.cell(row=current_transaction, column=2, value="-")
				wsdefault.cell(row=current_transaction, column=3, value="Error en connectar amb la maquina")
				wsdefault.cell(row=current_transaction, column=6, value="-")
				current_transaction += 1
				workbook.save(fitxer)
			print()
			i += 1
		Data("w")	#2.592.000			#escriure la ultima data aixis sap desde on mirar les copies, per activar aixo primer he de fer que anexi a el fitxer on envia
	f.close()
	return(Backups)
	f.close()

def statusConvertor(status):
	if status == 2:
		return("Correcte")
	elif status == 4:
		return("Warning")
	elif status == 5:
		return("ERROR")
	else:
		return("codi desconegut")

def tamanyRestant(i):
	with open('data/dispositius.json', 'r') as f:
		rconf = json.load(f)
		url2 = rconf['dispositius'][i]['url']+"webapi/entry.cgi"
		nom = rconf['dispositius'][i]['nom']
		cookie = rconf['dispositius'][i]['cookie']

		#aconseguir un apartat de la url (aut.cgi o entry.cgi)
		cookie = rconf['dispositius'][i]['cookie']
		query_parameters = {"api":"SYNO.API.Info", "version":"1", "method":"query", "query":"all"}
		queryUrl = rconf['dispositius'][i]['url']+"webapi/query.cgi"
		try:
			query = requests.get(queryUrl, params=query_parameters, headers={"cookie": cookie}).json()
			path= str(query['data']['SYNO.API.Auth']['path'])
		except Exception as e:
			print("Error en la query")
			now = datetime.datetime.now()
			date_string = now.strftime('%Y-%m-%d--%H-%M-%S-Query')
			a = open("errorLogs/"+date_string+".txt",'w')
			a.write(str(e))
			a.close()

		user = rconf['dispositius'][i]['user']
		password = rconf['dispositius'][i]['password']
		url = rconf['dispositius'][i]['url']+"webapi/"+path
		
		my_headers = {"cookie": cookie}
		try:
			sid = login(user, password, url, cookie)
			tamany_parameters = {"api":"SYNO.FileStation.List", "version":"2", "method":"list_share", "additional":'["volume_status"]', "_sid": sid}
			response = requests.get(url2, params=tamany_parameters, headers=my_headers).json()
			espaiLliure = round(((response['data']['shares'][0]['additional']['volume_status']['freespace']/1024)/1024)/1024)
			f.close()
			return(espaiLliure)
		except Exception as e:
			now = datetime.datetime.now()
			date_string = now.strftime('%Y-%m-%d--%H-%M-%S-Conexio')
			f = open("errorLogs/"+date_string+".txt",'w')
			f.write("Error en connectar amb la maquina "+nom)
			f.close()
			return("Fallo en la conexio")
	f.close()

# y es cada transaccio (es reseteja per cada dispositiu)
# z es personalitzat que es per cada dispositiu que tingui transaccio (es reseteja per cada NAS)
# x es cada dispositiu (es reseteja per cada NAS)
# i es cada nas (es reseteja cada execucio)
# current_transaction es cada transaccio (es reseteja cada execucio)

def escriureDades(nom_dispositiu, ws, status, temps_finalitzacio, tamany_transferencia, y, z, nom_nas, tamanyLliure):
	global current_transaction
	file_time = datetime.datetime.fromtimestamp(temps_finalitzacio)
	num1=4*z+1
	num2=4*z+2
	num3=4*z+3

	d = ws.cell(row=num1, column=1, value=nom_dispositiu)
	d = ws.cell(row=num2, column=1, value="Status")
	d = ws.cell(row=num3, column=1, value="Tamany MB")
	d = ws.cell(row=num1, column=y+2, value=file_time.strftime('%Y-%m-%d'))
	d = ws.cell(row=num2, column=y+2, value=status)
	d = ws.cell(row=num3, column=y+2, value=round((tamany_transferencia/1024)/1024))

	wsdefault = workbook['Sheet']

	if y==0:
		
		wsdefault.cell(row=current_transaction, column=1, value=nom_nas)
		wsdefault.cell(row=current_transaction, column=2, value=nom_dispositiu)
		wsdefault.cell(row=current_transaction, column=3, value=('=LOOKUP(2,1/('+nom_nas+'!'+str(num1)+':'+str(num1)+'<>""),'+nom_nas+'!'+str(num1)+':'+str(num1)+')'))
		wsdefault.cell(row=current_transaction, column=4, value=('=LOOKUP(2,1/('+nom_nas+'!'+str(num3)+':'+str(num3)+'<>""),'+nom_nas+'!'+str(num3)+':'+str(num3)+')'))
		wsdefault.cell(row=current_transaction, column=5, value=('=LOOKUP(2,1/('+nom_nas+'!'+str(num2)+':'+str(num2)+'<>""),'+nom_nas+'!'+str(num2)+':'+str(num2)+')'))
		wsdefault.cell(row=current_transaction, column=6, value=tamanyLliure)
		current_transaction += 1

def escriptorExcel(nom_dispositiu, status, temps_finalitzacio, tamany_transferencia, workbook, y, z, nom_nas, tamanyLliure):
	trobat = False
	for sheet in workbook:
		if sheet.title == nom_nas:
			trobat = True
			escriureDades(nom_dispositiu, sheet, status, temps_finalitzacio, tamany_transferencia, y, z, nom_nas, tamanyLliure)
			break
	if trobat == False:
		ws = workbook.create_sheet(nom_nas)
		escriureDades(nom_dispositiu, ws, status, temps_finalitzacio, tamany_transferencia, y, z, nom_nas, tamanyLliure)

def prepExcel(workbook):
	wsdefault = workbook['Sheet']
	wsdefault.cell(row=1, column=1, value="Nom NAS")
	wsdefault.cell(row=1, column=2, value="Nom Dispositiu")
	wsdefault.cell(row=1, column=3, value="Data")
	wsdefault.cell(row=1, column=4, value="Tamany MB")
	wsdefault.cell(row=1, column=5, value="Status")
	wsdefault.cell(row=1, column=6, value="Tamany Lliure GB")


if exists(fitxer) == False:
	workbook = Workbook()
	prepExcel(workbook)
	workbook.save(fitxer)
	
workbook = load_workbook(filename = fitxer)

for sheet in workbook:
	if sheet.title != "Sheet":
		workbook.remove(sheet)

llistaTransf = []
llistadispCopia = []
llistaNAS = []
llistaFinal = [] 
dadesCopiesTotes = recoleccioDades(workbook)
num_nas = len(dadesCopiesTotes)
i=0
nom_dispositiu=""
while i < num_nas:
	with open('data/dispositius.json', 'r') as f:
		rconf = json.load(f)
		nom_nas = rconf['dispositius'][i]['nom']
		id_pandora = rconf['dispositius'][i]['pandoraID']
		f.close()
	num_copies = int(dadesCopiesTotes[i]['data']['total'])
	tamanyLliure=tamanyRestant(i)
	x = 0
	z = 0
	while x < num_copies: 
		if len(dadesCopiesTotes[i]['data']['device_list'][x]['transfer_list']) != 0:
			num_transferencies = len(dadesCopiesTotes[i]['data']['device_list'][x]['transfer_list'])

			y=0
			while y < num_transferencies:
				nom_dispositiu = dadesCopiesTotes[i]['data']['device_list'][x]['transfer_list'][y]['device_name']
				status = statusConvertor(dadesCopiesTotes[i]['data']['device_list'][x]['transfer_list'][y]['status'])
				tamany_transferencia = dadesCopiesTotes[i]['data']['device_list'][x]['transfer_list'][y]['transfered_bytes']
				temps_finalitzacio = dadesCopiesTotes[i]['data']['device_list'][x]['transfer_list'][y]['time_end']
				if y==0:
					print()
					print(nom_dispositiu)
					print("[", end="")
				if status == "Correcte":
					print("#", end="")
				else:
					print("X", end="")

				file_time = datetime.datetime.fromtimestamp(temps_finalitzacio)
				dataF=file_time.strftime('%Y-%m-%d')
				if y+1 < num_transferencies:
					u=1
				else:
					llistaTransf.append({"data":dataF, "status":status, "tamany_transferenciaMB":(tamany_transferencia/1024)/1024})

				escriptorExcel(nom_dispositiu, status, temps_finalitzacio, tamany_transferencia, workbook, y, z, nom_nas, tamanyLliure)
				try:
					workbook.save(fitxer)
				except:
					now = datetime.datetime.now()
					date_string = now.strftime('%Y-%m-%d--%H-%M-%S-permisos')
					f = open("errorLogs/"+date_string+".txt",'w')
					f.write("Error de permisos en obrir el Excel (Pot ser que el excel estigui obert?)")
					f.close()
					print("Error de permisos")

				y += 1
			print("]", end="")
			print()
			z += 1
		llistadispCopia.append({"nomDispositiu":nom_dispositiu, "Transferencies":llistaTransf})
		nom_dispositiu = ""
		llistaTransf = []
		x += 1
	llistaNAS.append({"nomNAS":nom_nas, "ID Pandora":id_pandora, "copies":llistadispCopia})
	llistadispCopia = []
	i += 1
llistaFinal = [{"NAS":llistaNAS}]
if exists("dadesSynology.json") == True:
        os.remove("dadesSynology.json")
try:
	with open("dadesSynology.json", 'w') as f:
		json.dump(llistaFinal, f, indent = 4)
except Exception as e:
		print("Error d'escriptura de json")
		now = datetime.datetime.now()
		date_string = now.strftime('%Y-%m-%d--%H-%M-%S-json')
		f = open("errorLogs/"+date_string+".txt",'w')
		f.write("Error d'escriptura de json "+str(e))
		f.close()
