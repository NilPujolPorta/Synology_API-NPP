import requests
import json
from datetime import datetime,timezone
import datetime
import time
from os.path import exists	
import openpyxl				#pip install openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import os
import argparse
import mysql.connector
from configparser import ConfigParser

#millores a fer:
# apendre a anexar al excel per poder fer algunes funcions actualment commentades


parser = argparse.ArgumentParser(description='Una API per a recullir invormacio de varis NAS Synology que tinguin la versio 6 o mes.', epilog="Per configuracio adicional anar a config/api.conf")
parser.add_argument('-e', '--excel', help='Guardar la informacio a un excel, per defecte esta desactivat', action="store_true")
parser.add_argument('-q', '--quiet', help='Nomes mostra els errors i el missatge de acabada per pantalla.', action="store_false")
parser.add_argument('-f', '--file', help='Especificar el fitxer de excel a on guardar. Per defecte es revisio_copies_seguretat_synology_vs1.xlsx', default="revisio_copies_seguretat_synology_vs1.xlsx", metavar="RUTA")
parser.add_argument('-v', '--versio', help='Mostra la versio', action='version', version='Synology_API-NPP vs1.6.3')
args = parser.parse_args()

current_transaction = 2
fitxer = args.file 

def Data(WoR):
	if WoR == "w":
		with open('data/data.txt', 'w') as f:
			wdata = temps()-2592000
			f.write(str(wdata))
			f.close()
	elif WoR == "r":
		with open('data/data.txt', 'r') as f:
			rdata = f.read()
			f.close()
			return(rdata)
	else:
		print("Error en modificar data/data.txt (el metode de interaccio amb el fitxer es erroni o inexistent)")

def temps():
	dt = datetime.datetime.now(timezone.utc)
	utc_time = dt.replace(tzinfo=timezone.utc)
	utc_timestamp = utc_time.timestamp()
	return(round(utc_timestamp))


def login(user, password, url, cookie):
	login_parameters = {"api":"SYNO.API.Auth", "version":"3", "method":"login", "account": user, "passwd": password, "session":"ActiveBackup", "format":"cookie"}
	my_headers = {"cookie": cookie}
	response = requests.get(url, params=login_parameters, headers=my_headers).json()
	sid = response['data']['sid']
	
	if	response['success'] == True:
		if args.quiet:
			print("Login correcte")
	else:
		print("Login erroni")
		now = datetime.datetime.now()
		date_string = now.strftime('%Y-%m-%d--%H-%M-%S-login')
		f = open("errorLogs/"+date_string+".txt",'w')
		f.write(str(response))
		f.close()
		print(response)
	return(sid)

def logout(url, sid, cookie):
	logout_parameters = {"api":"SYNO.API.Auth", "version":"2", "method":"logout", "session":"ActiveBackup"}
	my_headers={"cookie": cookie}
	response = requests.get(url, params=logout_parameters, headers=my_headers).json()
	if	response['success'] == True:
		if args.quiet:
			print("Logout correcte")
		return(response)
	else:
		print("Logout erroni")
		now = datetime.datetime.now()
		date_string = now.strftime('%Y-%m-%d--%H-%M-%S-Logout')
		f = open("errorLogs/"+date_string+".txt",'w')
		f.write(str(response))
		f.close()
		print(response)

def InfoCopies(url, cookie, sid):
	copies_parameters = {"api":"SYNO.ActiveBackup.Overview", "version":"1", "method":"list_device_transfer_size", "time_start": int(Data("r")), "time_end": temps(), "_sid": sid}
	response = requests.get(url, params=copies_parameters, headers={"cookie":cookie}).json()
	if	response['success'] == True:
		if args.quiet:
			print("Operacio de dades de backup correcte")
		return(response)
	else:
		print("Operacio de dades de backup erroni")
		now = datetime.datetime.now()
		date_string = now.strftime('%Y-%m-%d--%H-%M-%S-Backups')
		f = open("errorLogs/"+date_string+".txt",'w')
		f.write(str(response))
		f.close()
		print(response)

def recoleccioDades(workbook):
	global current_transaction
	global fitxer
	Backups = []
	num_dispositius = len(taulabd)
	for x in taulabd:
		#aconseguir un apartat de la url (aut.cgi o entry.cgi)
		cookie = x[4]
		query_parameters = {"api":"SYNO.API.Info", "version":"1", "method":"query", "query":"all"}
		queryUrl = x[3]+"webapi/query.cgi"
		try:
			query = requests.get(queryUrl, params=query_parameters, headers={"cookie": cookie}).json()
			path= str(query['data']['SYNO.API.Auth']['path'])
		except:
			print("Error en la query")
			now = datetime.datetime.now()
			date_string = now.strftime('%Y-%m-%d--%H-%M-%S-Query')
			a = open("errorLogs/"+date_string+".txt",'w')
			a.write(str(query))
			a.close()
		user = x[1]
		password = x[2]
		url = x[3]+"webapi/"+path
		url2 = x[3]+"webapi/entry.cgi"
		nom = x[0]
		if args.quiet:
			print(nom)
		try:
			sid = login(user, password, url, cookie)
			Backups.append(InfoCopies(url2, cookie, sid))
			logout(url, sid, cookie)
		except:
			now = datetime.datetime.now()
			date_string = now.strftime('%Y-%m-%d--%H-%M-%S-Conexio')
			f = open("errorLogs/"+date_string+".txt",'w')
			f.write("Error en connectar amb la maquina "+nom)
			f.close()
			print("Error en connectar amb la maquina")
			wsdefault = workbook['Sheet']
			wsdefault.cell(row=current_transaction, column=1, value=nom)
			wsdefault.cell(row=current_transaction, column=2, value="-")
			wsdefault.cell(row=current_transaction, column=3, value="Error en connectar amb la maquina")
			wsdefault.cell(row=current_transaction, column=6, value="-")
			current_transaction += 1
			workbook.save(fitxer)
		if args.quiet:
			print()
	Data("w")	#2.592.000			#escriure la ultima data aixis sap desde on mirar les copies, per activar aixo primer he de fer que anexi a el fitxer on envia
	return(Backups)

def statusConvertor(status):
	if status == 2:
		return("Correcte")
	elif status == 4:
		return("Warning")
	elif status == 5:
		return("ERROR")
	else:
		return("codi desconegut")

def tamanyRestant(i):
	url2 = taulabd[i][3]+"webapi/entry.cgi"
	nom = taulabd[i][0]
	cookie = taulabd[i][4]
	#aconseguir un apartat de la url (aut.cgi o entry.cgi)
	queryUrl = taulabd[i][3]+"webapi/query.cgi"
	query_parameters = {"api":"SYNO.API.Info", "version":"1", "method":"query", "query":"all"}
	try:
		query = requests.get(queryUrl, params=query_parameters, headers={"cookie": cookie}).json()
		path= str(query['data']['SYNO.API.Auth']['path'])
		url = taulabd[i][3]+"webapi/"+path
	except Exception as e:
		print("Error en la query")
		now = datetime.datetime.now()
		date_string = now.strftime('%Y-%m-%d--%H-%M-%S-Query')
		a = open("errorLogs/"+date_string+".txt",'w')
		a.write(str(e))
		a.close()
	user = taulabd[i][1]
	password = taulabd[i][2]
	my_headers = {"cookie": cookie}
	try:
		sid = login(user, password, url, cookie)
		tamany_parameters = {"api":"SYNO.FileStation.List", "version":"2", "method":"list_share", "additional":'["volume_status"]', "_sid": sid}
		response = requests.get(url2, params=tamany_parameters, headers=my_headers).json()
		espaiLliure = round(((response['data']['shares'][0]['additional']['volume_status']['freespace']/1024)/1024)/1024)
		return(espaiLliure)
	except Exception as e:
		now = datetime.datetime.now()
		date_string = now.strftime('%Y-%m-%d--%H-%M-%S-Conexio')
		f = open("errorLogs/"+date_string+".txt",'w')
		f.write("Error en connectar amb la maquina "+ nom)
		f.close()
		return("Fallo en la conexio")

# y es cada transaccio (es reseteja per cada dispositiu)
# z es personalitzat que es per cada dispositiu que tingui transaccio (es reseteja per cada NAS)
# x es cada dispositiu (es reseteja per cada NAS)
# i es cada nas (es reseteja cada execucio)
# current_transaction es cada transaccio (es reseteja cada execucio)

def escriureDades(nom_dispositiu, ws, status, temps_finalitzacio, tamany_transferencia, y, z, nom_nas, tamanyLliure):
	global current_transaction
	file_time = datetime.datetime.fromtimestamp(temps_finalitzacio)
	num1=4*z+1
	num2=4*z+2
	num3=4*z+3

	d = ws.cell(row=num1, column=1, value=nom_dispositiu)
	d = ws.cell(row=num2, column=1, value="Status")
	d = ws.cell(row=num3, column=1, value="Tamany MB")
	d = ws.cell(row=num1, column=y+2, value=file_time.strftime('%Y-%m-%d'))
	d = ws.cell(row=num2, column=y+2, value=status)
	d = ws.cell(row=num3, column=y+2, value=round((tamany_transferencia/1024)/1024))

	wsdefault = workbook['Sheet']

	if y==0:
		
		wsdefault.cell(row=current_transaction, column=1, value=nom_nas)
		wsdefault.cell(row=current_transaction, column=2, value=nom_dispositiu)
		wsdefault.cell(row=current_transaction, column=3, value=('=LOOKUP(2,1/('+nom_nas+'!'+str(num1)+':'+str(num1)+'<>""),'+nom_nas+'!'+str(num1)+':'+str(num1)+')'))
		wsdefault.cell(row=current_transaction, column=4, value=('=LOOKUP(2,1/('+nom_nas+'!'+str(num3)+':'+str(num3)+'<>""),'+nom_nas+'!'+str(num3)+':'+str(num3)+')'))
		wsdefault.cell(row=current_transaction, column=5, value=('=LOOKUP(2,1/('+nom_nas+'!'+str(num2)+':'+str(num2)+'<>""),'+nom_nas+'!'+str(num2)+':'+str(num2)+')'))
		wsdefault.cell(row=current_transaction, column=6, value=tamanyLliure)
		current_transaction += 1

def escriptorExcel(nom_dispositiu, status, temps_finalitzacio, tamany_transferencia, workbook, y, z, nom_nas, tamanyLliure):
	trobat = False
	for sheet in workbook:
		if sheet.title == nom_nas:
			trobat = True
			escriureDades(nom_dispositiu, sheet, status, temps_finalitzacio, tamany_transferencia, y, z, nom_nas, tamanyLliure)
			break
	if trobat == False:
		ws = workbook.create_sheet(nom_nas)
		escriureDades(nom_dispositiu, ws, status, temps_finalitzacio, tamany_transferencia, y, z, nom_nas, tamanyLliure)

def prepExcel(workbook):
	wsdefault = workbook['Sheet']
	wsdefault.cell(row=1, column=1, value="Nom NAS")
	wsdefault.cell(row=1, column=2, value="Nom Dispositiu")
	wsdefault.cell(row=1, column=3, value="Data")
	wsdefault.cell(row=1, column=4, value="Tamany MB")
	wsdefault.cell(row=1, column=5, value="Status")
	wsdefault.cell(row=1, column=6, value="Tamany Lliure GB")
###################################################################################################################################################################
if exists("config/config.ini"):
	configuracio = True
else:
	print("Emplena el fitxer de configuracio de Base de Dades a config/config.ini")
	config_object = ConfigParser()
	config_object["SRV_BDD"] = {
    	"host": "localhost",
    	"user": "root",
    	"passwd": "patata"
	}
	with open('config.ini', 'w') as conf:
		config_object.write(conf)

config_object = ConfigParser()
config_object.read("config.ini")
srvinfo = config_object["SRV_BDD"]
servidor = "{}".format(srvinfo["host"])
usuari = "{}".format(srvinfo["user"])
contrassenya = "{}".format(srvinfo["passwd"])

try:
    mydb =mysql.connector.connect(
        host=servidor,
        user=usuari,
        password=contrassenya,
        database="synology"
        )
    mycursor = mydb.cursor(buffered=True)
    print("Access BDD correcte")
except:
	try:
        
		mydb =mysql.connector.connect(
            host=servidor,
            user=usuari,
            password=contrassenya
            )
		print("Base de dades no existeix, creant-la ...")
		mycursor = mydb.cursor(buffered=True)
		mycursor.execute("CREATE DATABASE synology")
		mydb =mysql.connector.connect(
            host=servidor,
            user=usuari,
            password=contrassenya,
            database="synology"
            )
		mycursor = mydb.cursor(buffered=True)
		mycursor.execute("CREATE TABLE dispositius (nom VARCHAR(255), usuari VARCHAR(255), contassenya VARCHAR(255), url VARCHAR(255), cookie VARCHAR(400), pandoraID INT(3));")
	except:
		print("Login BDD incorrecte")
		quit()
taulabd = []

mycursor.execute("SELECT * FROM dispositius")
resultatbd = mycursor.fetchall()
for fila in resultatbd:
	taulabd.append(fila)

if exists(fitxer) == False:
	workbook = Workbook()
	prepExcel(workbook)
	workbook.save(fitxer)
	
workbook = load_workbook(filename = fitxer)

for sheet in workbook:
	if sheet.title != "Sheet":
		workbook.remove(sheet)

llistaTransf = []
llistadispCopia = []
llistaNAS = []
llistaFinal = [] 
dadesCopiesTotes = recoleccioDades(workbook)
num_nas = len(dadesCopiesTotes)
i=0
nom_dispositiu=""

while i < num_nas:
	nom_nas = taulabd[i][0]
	id_pandora = taulabd[i][5]
	num_copies = int(dadesCopiesTotes[i]['data']['total'])
	tamanyLliure=tamanyRestant(i)
	x = 0
	z = 0
	while x < num_copies: 
		if len(dadesCopiesTotes[i]['data']['device_list'][x]['transfer_list']) != 0:
			num_transferencies = len(dadesCopiesTotes[i]['data']['device_list'][x]['transfer_list'])

			y=0
			while y < num_transferencies:
				nom_dispositiu = dadesCopiesTotes[i]['data']['device_list'][x]['transfer_list'][y]['device_name']
				status = statusConvertor(dadesCopiesTotes[i]['data']['device_list'][x]['transfer_list'][y]['status'])
				tamany_transferencia = dadesCopiesTotes[i]['data']['device_list'][x]['transfer_list'][y]['transfered_bytes']
				temps_finalitzacio = dadesCopiesTotes[i]['data']['device_list'][x]['transfer_list'][y]['time_end']
				if y==0:
					if args.quiet:
						print()
						print(nom_dispositiu)
						print("[", end="")
				if status == "Correcte":
					if args.quiet:
						print("#", end="")
				elif args.quiet:
					print("X", end="")

				file_time = datetime.datetime.fromtimestamp(temps_finalitzacio)
				dataF=file_time.strftime('%Y-%m-%d')
				if y+1 < num_transferencies:
					u=1
				else:
					llistaTransf.append({"data":dataF, "status":status, "tamany_transferenciaMB":(tamany_transferencia/1024)/1024})
				if args.excel:
					escriptorExcel(nom_dispositiu, status, temps_finalitzacio, tamany_transferencia, workbook, y, z, nom_nas, tamanyLliure)

					try:
						workbook.save(fitxer)
					except:
						now = datetime.datetime.now()
						date_string = now.strftime('%Y-%m-%d--%H-%M-%S-permisos')
						f = open("errorLogs/"+date_string+".txt",'w')
						f.write("Error de permisos en obrir el Excel (Pot ser que el excel estigui obert?)")
						f.close()
						print("Error de permisos")

				y += 1
			if args.quiet:
				print("]", end="")
				print()
			z += 1
		llistadispCopia.append({"nomDispositiu":nom_dispositiu, "Transferencies":llistaTransf})
		nom_dispositiu = ""
		llistaTransf = []
		x += 1
	llistaNAS.append({"nomNAS":nom_nas, "ID Pandora":id_pandora, "copies":llistadispCopia})
	llistadispCopia = []
	i += 1
llistaFinal = [{"NAS":llistaNAS}]
if exists("dadesSynology.json") == True:
        os.remove("dadesSynology.json")
try:
	with open("dadesSynology.json", 'w') as f:
		json.dump(llistaFinal, f, indent = 4)
except Exception as e:
		print("Error d'escriptura de json")
		now = datetime.datetime.now()
		date_string = now.strftime('%Y-%m-%d--%H-%M-%S-json')
		f = open("errorLogs/"+date_string+".txt",'w')
		f.write("Error d'escriptura de json "+str(e))
		f.close()
if not(args.quiet):
	print("Done")